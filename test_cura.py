import openpyxl
import pytest
from selenium.webdriver.support.select import Select
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
import time

service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)
driver.get("https://katalon-demo-cura.herokuapp.com")
driver.maximize_window()
time.sleep(3)



path = "cura.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

user_name = sheet_obj.cell(row=1, column=1).value
password = sheet_obj.cell(row=1, column=2).value
my_comment = sheet_obj.cell(row=1, column=3).value
print(user_name, password, my_comment)




make_appoinment = driver.find_element(By.XPATH, "/html[1]/body[1]/header[1]/div[1]/a[1]")
make_appoinment.click()
time.sleep(2)

input_usernaem =  driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/div[2]/form[1]/div[2]/div[1]/input[1]")
input_usernaem.send_keys(user_name)
input_password = driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/div[2]/form[1]/div[3]/div[1]/input[1]")
input_password.send_keys(password + Keys.ENTER)
time.sleep(1)

select_facility = driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/form[1]/div[1]/div[1]/select[1]")
select_facility.click()
time.sleep(1)

my_select_facility = Select(select_facility);
my_select_facility.select_by_visible_text("Seoul CURA Healthcare Center")
time.sleep(1)

apply_hospital_readmission = driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/form[1]/div[2]/div[1]/label[1]/input[1]")
apply_hospital_readmission.click()
time.sleep(1)

select_program = driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/form[1]/div[3]/div[1]/label[2]/input[1]")
select_program.click()
time.sleep(1)

select_calendar = driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/form[1]/div[4]/div[1]/div[1]/div[1]/span[1]")
select_calendar.click()
time.sleep(1)

input_select_calendar = driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/form[1]/div[4]/div[1]/div[1]/input[1]")
input_select_calendar.click()
input_select_calendar.send_keys("06/10/2002")
time.sleep(1)

input_comment = driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/form[1]/div[5]/div[1]/textarea[1]")
input_comment.click()
input_comment.send_keys(my_comment)
time.sleep(1)

input_book_appointment = driver.find_element(By.XPATH, "/html[1]/body[1]/section[1]/div[1]/div[1]/form[1]/div[6]/div[1]/button[1]")
input_book_appointment.click()
time.sleep(2)

actual_title = driver.find_element(By.XPATH, "//*[text()='Appointment Confirmation']")

def case1(actual_title):
    return bool(actual_title)

def test_case1():
    assert case1("Appointment Confirmation") == True
